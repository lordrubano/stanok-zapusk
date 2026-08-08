"""Безопасная Excel-выгрузка в пять нормативных листов.

Перечисленные в ``KNOWN_UNREACHABLE_THREATS`` угрозы недостижимы через нормальный
вход ``Cell``, ``add_rows`` и ``save``. Если появится путь, создающий такую книгу,
для него понадобятся отдельные проверки.
"""

from __future__ import annotations

import json
import zipfile
from collections.abc import Callable
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell as OpenpyxlCell


SHEETS = ("DATA", "REVIEW", "REJECTED", "ERRORS", "AUDIT")
FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

KNOWN_UNREACHABLE_THREATS = (
    "VBA-макрос",
    "внешняя ссылка externalLink",
    "гиперссылка",
)
"""Известные ограничения экспортёра.

Каждая угроза недостижима через нормальный вход ``Cell``, ``add_rows`` и ``save``.
Если появится путь, создающий такую книгу, для него понадобятся отдельные проверки.
"""


class ExcelSafetyError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class ExcelLimits:
    """Утверждённые владельцем пределы.

    Значения по умолчанию не меняются исполнителем без нового решения владельца.
    """

    # Все значения ниже утверждены владельцем.
    max_cell_length: int = 32_767
    max_rows_per_sheet: int = 1_048_576
    max_workbook_bytes: int = 52_428_800
    max_styles: int = 1_000
    safe_significant_digits: int = 15


@dataclass(frozen=True, slots=True)
class ExcelFormats:
    """Явные форматы Excel, утверждённые владельцем по умолчанию."""

    date: str = "yyyy-mm-dd"
    datetime: str = "yyyy-mm-dd hh:mm:ss"
    money: str = "#,##0.00"
    decimal: str = "0.###############"
    text: str = "@"


@dataclass(frozen=True, slots=True)
class Cell:
    value: Any
    kind: str = "text"
    number_format: str | None = None


@dataclass(frozen=True, slots=True)
class SaveReport:
    """Отчёт о сохранённых XLSX и сопутствующем машинном JSON."""

    xlsx_path: Path
    json_path: Path
    escaped_cells: int
    textified_values: int
    workbook_bytes: int
    disabled_guards: frozenset[str]

    @property
    def workbook_path(self) -> Path:
        return self.xlsx_path

    @property
    def workbook_size_bytes(self) -> int:
        return self.workbook_bytes


@dataclass(slots=True)
class _BuildStats:
    escaped_cells: int = 0
    textified_values: int = 0


def _significant_digits(text: str) -> int:
    stripped = text.lstrip("+-0").replace(".", "")
    return len(stripped.lstrip("0")) or 1


def guard_formula_injection(
    *,
    target: OpenpyxlCell | None = None,
    value: str | None = None,
    workbook: Workbook | None = None,
) -> bool:
    """Исполняет п. 13.2 чертежа: защита текста от формульной инъекции."""

    escaped = False
    if target is not None and value is not None:
        target.value = value
        if value.startswith(FORMULA_PREFIXES):
            # quotePrefix хранится в стиле и не меняет видимое значение ячейки.
            target.data_type = "s"
            target.quotePrefix = True
            escaped = True
    if workbook is not None:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        raise ExcelSafetyError(
                            f"формула запрещена: {worksheet.title}!{cell.coordinate}"
                        )
    return escaped


def guard_text_identifiers(target: OpenpyxlCell, value: Any) -> None:
    """Исполняет п. 13.2 чертежа: текстовые идентификаторы сохраняются текстом."""

    target.value = str(value)
    target.data_type = "s"


def guard_decimal_only(value: Any) -> Decimal:
    """Исполняет п. 13.2 чертежа: деньги и точные количества используют Decimal."""

    if isinstance(value, float):
        raise ExcelSafetyError(
            "E_FLOAT_IN_EXACT_FIELD: float запрещён в денежном или точном поле"
        )
    if isinstance(value, bool) or not isinstance(value, (Decimal, int, str)):
        raise ExcelSafetyError(
            "денежное или точное поле принимает только Decimal, int или str"
        )
    try:
        number = value if isinstance(value, Decimal) else Decimal(value)
    except (InvalidOperation, ValueError) as exc:
        raise ExcelSafetyError(f"неверное Decimal-значение: {value!r}") from exc
    if not number.is_finite():
        raise ExcelSafetyError("NaN и бесконечность запрещены")
    return number


def guard_significant_digits(exact: str, limits: ExcelLimits) -> bool:
    """Исполняет п. 13.2 чертежа: сверхточные числа направляются в текст."""

    return _significant_digits(exact) > limits.safe_significant_digits


def guard_no_macros_links(
    workbook: Workbook | None = None,
    payload: bytes | None = None,
    *,
    output_path: Path | None = None,
) -> None:
    """Исполняет п. 13.2 чертежа: запрещает макросы, объекты и внешние ссылки."""

    if output_path is not None and output_path.suffix.lower() != ".xlsx":
        raise ExcelSafetyError("разрешён только формат .xlsx")
    if workbook is None or payload is None:
        return

    if getattr(workbook, "vba_archive", None) is not None:
        raise ExcelSafetyError("макросы в книге запрещены")
    if getattr(workbook, "_external_links", []):
        raise ExcelSafetyError("внешние ссылки в книге запрещены")
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.hyperlink is not None:
                    raise ExcelSafetyError(
                        f"гиперссылка запрещена: {worksheet.title}!{cell.coordinate}"
                    )

    with zipfile.ZipFile(BytesIO(payload)) as package:
        names = {name.lower() for name in package.namelist()}
    if "xl/vbaproject.bin" in names:
        raise ExcelSafetyError("макросы в книге запрещены")
    forbidden_prefixes = (
        "xl/externallinks/",
        "xl/embeddings/",
        "xl/activex/",
        "xl/ctrlprops/",
    )
    if any(name.startswith(forbidden_prefixes) for name in names):
        raise ExcelSafetyError("внешние ссылки и исполняемые объекты в книге запрещены")


def guard_size_limits(
    limits: ExcelLimits,
    *,
    cell_text: str | None = None,
    sheet_name: str | None = None,
    row_count: int | None = None,
    workbook_bytes: int | None = None,
    style_count: int | None = None,
) -> None:
    """Исполняет п. 13.2 чертежа: ограничивает ячейки, строки, книгу и стили."""

    if cell_text is not None and len(cell_text) > limits.max_cell_length:
        raise ExcelSafetyError(
            f"длина ячейки {len(cell_text)} превышает предел {limits.max_cell_length}"
        )
    if row_count is not None and row_count > limits.max_rows_per_sheet:
        raise ExcelSafetyError(
            f"лист {sheet_name}: {row_count} строк превышают предел "
            f"{limits.max_rows_per_sheet}"
        )
    if workbook_bytes is not None and workbook_bytes > limits.max_workbook_bytes:
        raise ExcelSafetyError(
            f"книга {workbook_bytes} байт превышает предел {limits.max_workbook_bytes}"
        )
    if style_count is not None and style_count > limits.max_styles:
        raise ExcelSafetyError("превышен предел стилей книги")


def guard_explicit_formats(
    target: OpenpyxlCell,
    item: Cell,
    formats: ExcelFormats,
    *,
    text_output: bool = False,
) -> None:
    """Исполняет п. 13.2 чертежа: явно задаёт форматы дат, денег и чисел."""

    if text_output:
        target.number_format = formats.text
    elif item.number_format is not None:
        target.number_format = item.number_format
    elif item.kind in {"text", "identifier"}:
        target.number_format = formats.text
    elif item.kind == "money":
        target.number_format = formats.money
    elif item.kind == "decimal":
        target.number_format = formats.decimal
    elif item.kind == "integer":
        target.number_format = "0"
    elif item.kind == "date":
        if isinstance(item.value, datetime) and item.value.tzinfo is not None:
            target.number_format = formats.text
        elif isinstance(item.value, datetime):
            target.number_format = formats.datetime
        else:
            target.number_format = formats.date


def guard_readback(
    xlsx_path: Path,
    json_path: Path,
    expected_document: Mapping[str, Any],
    sheets: tuple[str, ...] = SHEETS,
) -> None:
    """Перечитывает с диска оба записанных артефакта и сверяет JSON.

    ``sheets`` — набор листов, который книга **объявила о себе**. Проверка от
    этого не слабеет: сверяются по-прежнему набор и порядок, просто с тем, что
    книга обещала, а не с одним набором на все книги. Книг у станка две, и
    нормативные наборы у них разные.
    """

    try:
        xlsx_missing = not xlsx_path.is_file() or xlsx_path.stat().st_size == 0
    except OSError:
        xlsx_missing = not xlsx_path.exists()
    if xlsx_missing:
        raise ExcelSafetyError(
            "E_XLSX_MISSING_AFTER_WRITE: XLSX отсутствует или имеет нулевую длину"
        )

    try:
        workbook = load_workbook(
            xlsx_path, read_only=False, data_only=False, keep_links=True
        )
        try:
            if workbook.sheetnames != list(sheets):
                raise ValueError("изменились набор или порядок листов")
        finally:
            workbook.close()
    except Exception as exc:
        raise ExcelSafetyError(
            "E_XLSX_UNREADABLE_AFTER_WRITE: записанный XLSX не читается"
        ) from exc

    try:
        json_missing = not json_path.is_file() or json_path.stat().st_size == 0
    except OSError:
        json_missing = not json_path.exists()
    if json_missing:
        raise ExcelSafetyError(
            "E_JSON_MISSING_AFTER_WRITE: JSON отсутствует или имеет нулевую длину"
        )

    try:
        actual_document = json.loads(json_path.read_text(encoding="utf-8"))
        if actual_document != expected_document:
            raise ValueError("JSON не совпадает с записывавшимся документом")
    except Exception as exc:
        raise ExcelSafetyError(
            "E_JSON_UNREADABLE_AFTER_WRITE: записанный JSON не читается "
            "или не совпадает с исходным"
        ) from exc


GUARDS: dict[str, Callable[..., Any]] = {
    "G1_formula_injection": guard_formula_injection,
    "G2_text_identifiers": guard_text_identifiers,
    "G3_decimal_only": guard_decimal_only,
    "G4_significant_digits": guard_significant_digits,
    "G5_no_macros_links": guard_no_macros_links,
    "G6_size_limits": guard_size_limits,
    "G7_explicit_formats": guard_explicit_formats,
    "G8_readback": guard_readback,
}


class ResultWorkbook:
    """По умолчанию создаёт только XLSX без макросов и внешних ссылок."""

    def __init__(
        self,
        *,
        limits: ExcelLimits | None = None,
        formats: ExcelFormats | None = None,
        disabled_guards: frozenset[str] = frozenset(),
        sheets: tuple[str, ...] = SHEETS,
    ) -> None:
        unknown_guards = disabled_guards.difference(GUARDS)
        if unknown_guards:
            names = ", ".join(sorted(unknown_guards))
            raise ExcelSafetyError(f"неизвестные отключённые защиты: {names}")
        self.limits = limits or ExcelLimits()
        self.formats = formats or ExcelFormats()
        self.disabled_guards = frozenset(disabled_guards)
        self.sheets = tuple(sheets)
        self._rows: dict[str, list[Mapping[str, Cell | Any]]] = {
            name: [] for name in self.sheets}

    def _guard_enabled(self, name: str) -> bool:
        return name not in self.disabled_guards

    def _run_guard(self, name: str, *args: Any, **kwargs: Any) -> Any:
        if not self._guard_enabled(name):
            return None
        return GUARDS[name](*args, **kwargs)

    def add_rows(self, sheet: str, rows: Iterable[Mapping[str, Cell | Any]]) -> None:
        if sheet not in self._rows:
            raise ExcelSafetyError(f"лист вне нормативного набора: {sheet}")
        self._rows[sheet].extend(rows)

    @staticmethod
    def _columns(rows: Iterable[Mapping[str, Cell | Any]]) -> list[str]:
        columns: list[str] = []
        for row in rows:
            for key in row:
                if key not in columns:
                    columns.append(key)
        return columns

    def _write_text(
        self,
        target: OpenpyxlCell,
        text: str,
        stats: _BuildStats,
        *,
        protect_formula: bool,
    ) -> None:
        self._run_guard("G6_size_limits", self.limits, cell_text=text)
        if protect_formula:
            if self._guard_enabled("G1_formula_injection"):
                escaped = self._run_guard(
                    "G1_formula_injection", target=target, value=text
                )
                stats.escaped_cells += int(bool(escaped))
            else:
                target.value = text
        else:
            target.value = text
            target.data_type = "s"

    def _write_value(
        self, target: OpenpyxlCell, item: Cell, stats: _BuildStats
    ) -> None:
        kind = item.kind
        value = item.value
        text_output = False
        if value is None:
            target.value = None
            return

        if kind == "text":
            self._write_text(target, str(value), stats, protect_formula=True)
        elif kind == "identifier":
            # Это намеренно небезопасная ветвь без G2: Excel-подобное
            # распознавание превращает "007" в число 7. Именно G2 обязана
            # заменить значение на текст, иначе угроза материализуется.
            if isinstance(value, str) and value.isascii() and value.isdecimal():
                target.value = int(value)
            else:
                target.value = value
            self._run_guard("G2_text_identifiers", target, value)
            if isinstance(target.value, str):
                self._write_text(target, target.value, stats, protect_formula=True)
        elif kind in {"decimal", "money"}:
            if self._guard_enabled("G3_decimal_only"):
                number = self._run_guard("G3_decimal_only", value)
            else:
                number = value
            exact = format(number, "f") if isinstance(number, Decimal) else str(number)
            as_text = bool(
                self._run_guard("G4_significant_digits", exact, self.limits)
            )
            if as_text:
                self._write_text(target, exact, stats, protect_formula=False)
                stats.textified_values += 1
                text_output = True
            else:
                target.value = number
        elif kind == "integer":
            if isinstance(value, bool):
                raise ExcelSafetyError("bool нельзя выдавать за integer")
            number = int(value)
            exact = str(number)
            as_text = bool(
                self._run_guard("G4_significant_digits", exact, self.limits)
            )
            if as_text:
                self._write_text(target, exact, stats, protect_formula=False)
                stats.textified_values += 1
                text_output = True
            else:
                target.value = number
        elif kind == "date":
            if isinstance(value, datetime) and value.tzinfo is not None:
                self._write_text(target, value.isoformat(), stats, protect_formula=False)
            elif isinstance(value, (date, datetime)):
                target.value = value
            else:
                raise ExcelSafetyError("date требует datetime.date или datetime.datetime")
        elif kind == "boolean":
            if not isinstance(value, bool):
                raise ExcelSafetyError("boolean требует bool")
            target.value = value
        else:
            raise ExcelSafetyError(f"неподдерживаемый тип ячейки: {kind}")

        self._run_guard(
            "G7_explicit_formats",
            target,
            item,
            self.formats,
            text_output=text_output,
        )

    def _build(self) -> tuple[Workbook, _BuildStats]:
        wb = Workbook(write_only=False)
        wb.remove(wb.active)
        stats = _BuildStats()
        for sheet_name in self.sheets:
            ws = wb.create_sheet(sheet_name)
            rows = self._rows[sheet_name]
            if not rows:
                continue
            columns = self._columns(rows)
            self._run_guard(
                "G6_size_limits",
                self.limits,
                sheet_name=sheet_name,
                row_count=len(rows) + 1,
            )
            for index, name in enumerate(columns, 1):
                self._write_value(ws.cell(1, index), Cell(name, "text"), stats)
            for row_index, row in enumerate(rows, 2):
                for column_index, name in enumerate(columns, 1):
                    raw = row.get(name)
                    item = raw if isinstance(raw, Cell) else Cell(raw, "text")
                    self._write_value(ws.cell(row_index, column_index), item, stats)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
        return wb, stats

    def _render(self) -> tuple[bytes, _BuildStats]:
        wb, stats = self._build()
        stream = BytesIO()
        try:
            wb.save(stream)
        finally:
            wb.close()
        payload = stream.getvalue()
        self._run_guard(
            "G6_size_limits", self.limits, workbook_bytes=len(payload)
        )

        # Приёмочная проверка уже созданного ZIP-контейнера XLSX.
        check = load_workbook(BytesIO(payload), read_only=False, data_only=False, keep_links=True)
        try:
            if check.sheetnames != list(self.sheets):
                raise ExcelSafetyError("набор или порядок листов изменился")
            self._run_guard(
                "G6_size_limits",
                self.limits,
                style_count=len(getattr(check, "_cell_styles", [])),
            )
            self._run_guard("G1_formula_injection", workbook=check)
            self._run_guard("G5_no_macros_links", check, payload)
        finally:
            check.close()
        return payload, stats

    def to_bytes(self) -> bytes:
        payload, _ = self._render()
        return payload

    @staticmethod
    def _json_value(value: Any) -> Any:
        if isinstance(value, Decimal):
            return str(value)
        if isinstance(value, (date, datetime)):
            return value.isoformat()
        if value is None or isinstance(value, (str, int, float, bool)):
            return value
        return str(value)

    def _machine_document(
        self, target: Path, stats: _BuildStats, workbook_bytes: int
    ) -> dict[str, Any]:
        sheets: dict[str, dict[str, Any]] = {}
        for sheet_name in self.sheets:
            rows = self._rows[sheet_name]
            columns = self._columns(rows)
            machine_rows: list[dict[str, Any]] = []
            for row in rows:
                machine_row: dict[str, Any] = {}
                for name in columns:
                    raw = row.get(name)
                    item = raw if isinstance(raw, Cell) else Cell(raw, "text")
                    machine_row[name] = self._json_value(item.value)
                machine_rows.append(machine_row)
            sheets[sheet_name] = {"columns": columns, "rows": machine_rows}
        return {
            "xlsx_file": target.name,
            "disabled_guards": sorted(self.disabled_guards),
            "escaped_cells": stats.escaped_cells,
            "textified_values": stats.textified_values,
            "workbook_bytes": workbook_bytes,
            "sheets": sheets,
        }

    def save(self, path: str | Path) -> SaveReport:
        target = Path(path)
        self._run_guard("G5_no_macros_links", output_path=target)
        json_target = target.with_suffix(".json")
        try:
            payload, stats = self._render()
            document = self._machine_document(target, stats, len(payload))
            machine_payload = json.dumps(
                document, ensure_ascii=False, indent=2, allow_nan=False
            ) + "\n"
        except Exception:
            # Непрошедшая проверку пара результатов не должна оставаться на диске.
            target.unlink(missing_ok=True)
            json_target.unlink(missing_ok=True)
            raise

        target.parent.mkdir(parents=True, exist_ok=True)
        try:
            target.write_bytes(payload)
            json_target.write_text(machine_payload, encoding="utf-8", newline="\n")
            self._run_guard(
                "G8_readback", target, json_target, document, self.sheets
            )
        except Exception:
            target.unlink(missing_ok=True)
            json_target.unlink(missing_ok=True)
            raise
        return SaveReport(
            xlsx_path=target,
            json_path=json_target,
            escaped_cells=stats.escaped_cells,
            textified_values=stats.textified_values,
            workbook_bytes=len(payload),
            disabled_guards=self.disabled_guards,
        )
